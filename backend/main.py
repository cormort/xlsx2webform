#!/usr/bin/env python3
"""
Budget Table Editor Backend
FastAPI application for XLSX upload, table editing, and JSON import/export.
"""

import os
import io
import re
import csv
import glob
import uuid
import json
import hmac
import shutil
import secrets
import hashlib
import logging
import bcrypt
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
from urllib.parse import quote

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Header, Depends, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from backend.xlsx_parser import process_xlsx
from backend import registry

app = FastAPI(title="Budget Table Editor", version="1.0.0")

# Setup rate limiter
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS configuration
allowed_origins_raw = os.environ.get("ALLOWED_ORIGINS", "")
if allowed_origins_raw:
    origins = [o.strip() for o in allowed_origins_raw.split(",") if o.strip()]
else:
    origins = ["*"]

allow_credentials = True
if "*" in origins:
    allow_credentials = False

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=allow_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Paths whose GETs are polled/static — skipped by the audit log to cut noise.
_AUDIT_SKIP_GET = {
    "/api/auth/verify", "/api/admin/verify", "/api/registry",
    "/api/fund-names", "/api/supervisors",
}

# HTTP Security Headers + audit-log Middleware
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "frame-ancestors 'none';"
    )

    # Audit: record every API/fill action (skip polled & static GETs)
    path = request.url.path
    if (path.startswith("/api/") or path.startswith("/fill/")) and \
       not (request.method == "GET" and path in _AUDIT_SKIP_GET):
        try:
            auth = get_current_auth_optional(request.headers.get("X-Admin-Token"))
            client = request.client.host if request.client else ""
            _audit(f"{request.method} {path}", auth["role"], auth.get("email"),
                   client, detail=str(response.status_code))
        except Exception as e:
            logger.warning("audit failed: %s", e)

    return response

# Upload directory
UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
TEMPLATE_DIR = Path(__file__).parent.parent / "templates"
UPLOAD_DIR.mkdir(exist_ok=True)
TEMPLATE_DIR.mkdir(exist_ok=True)

DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

AUDIT_LOG = DATA_DIR / "audit.log"

# Logging
logger = logging.getLogger(__name__)


def _audit(action: str, role: str, email, ip: str, detail: str = ""):
    """Append one JSON line to the audit log. Never raises."""
    rec = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "role": role,
        "email": email or "",
        "ip": ip or "",
        "action": action,
        "detail": detail,
    }
    try:
        with AUDIT_LOG.open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception as e:
        logger.warning("audit write failed: %s", e)


def _read_audit(limit: int = 500, email: str = "", action: str = ""):
    """Return the most recent audit records (newest first), optionally filtered."""
    if not AUDIT_LOG.exists():
        return []
    email = (email or "").strip().lower()
    action = (action or "").strip().lower()
    out = []
    try:
        lines = AUDIT_LOG.read_text(encoding="utf-8").splitlines()
    except Exception as e:
        logger.warning("audit read failed: %s", e)
        return []
    for line in reversed(lines):
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except Exception:
            continue
        if email and email not in rec.get("email", "").lower():
            continue
        if action and action not in rec.get("action", "").lower():
            continue
        out.append(rec)
        if len(out) >= limit:
            break
    return out

# Upload size limit (10 MB)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# ── Admin & User Authentication ──────────────────────────
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
ADMIN_SECRET = os.environ.get("ADMIN_SECRET", secrets.token_hex(32))
ADMIN_TOKEN_EXPIRY_HOURS = int(os.environ.get("ADMIN_TOKEN_EXPIRY_HOURS", "24"))
admin_tokens: dict[str, datetime] = {}  # token → expiry datetime
users: dict[str, dict] = {}             # email → {password_hash, created_at}
user_tokens: dict[str, dict] = {}        # token → {email, expiry: datetime}


def _generate_admin_token() -> str:
    """Generate a signed admin token with expiry."""
    token_id = secrets.token_hex(24)
    expiry = datetime.now() + timedelta(hours=ADMIN_TOKEN_EXPIRY_HOURS)
    admin_tokens[token_id] = expiry
    # Persist admin tokens
    _save_admin_tokens()
    return token_id


def _validate_admin_token(token: str) -> bool:
    """Validate an admin token. Returns True if valid and not expired."""
    if not token or token not in admin_tokens:
        return False
    expiry = admin_tokens[token]
    if datetime.now() > expiry:
        del admin_tokens[token]
        _save_admin_tokens()
        return False
    return True


def _save_admin_tokens():
    """Persist valid admin tokens to disk."""
    data = {k: v.isoformat() for k, v in admin_tokens.items() if datetime.now() < v}
    (DATA_DIR / "admin_tokens.json").write_text(
        json.dumps(data, ensure_ascii=False), encoding='utf-8')


def _load_admin_tokens():
    """Load admin tokens from disk."""
    path = DATA_DIR / "admin_tokens.json"
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
            now = datetime.now()
            for token_id, expiry_str in data.items():
                expiry = datetime.fromisoformat(expiry_str)
                if now < expiry:
                    admin_tokens[token_id] = expiry
        except Exception as e:
            logger.warning("Failed to load admin tokens: %s", e)


def _save_users():
    """Persist users to disk."""
    (DATA_DIR / "users.json").write_text(
        json.dumps(users, ensure_ascii=False, indent=2), encoding='utf-8')


def _load_users():
    """Load users from disk."""
    path = DATA_DIR / "users.json"
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
            users.update(data)
        except Exception as e:
            logger.warning("Failed to load users: %s", e)


def _save_user_tokens():
    """Persist user tokens to disk."""
    data = {
        k: {"email": v["email"], "expiry": v["expiry"].isoformat()}
        for k, v in user_tokens.items()
        if datetime.now() < v["expiry"]
    }
    (DATA_DIR / "user_tokens.json").write_text(
        json.dumps(data, ensure_ascii=False), encoding='utf-8')


def _load_user_tokens():
    """Load user tokens from disk."""
    path = DATA_DIR / "user_tokens.json"
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
            now = datetime.now()
            for token_id, val in data.items():
                expiry = datetime.fromisoformat(val["expiry"])
                if now < expiry:
                    user_tokens[token_id] = {"email": val["email"], "expiry": expiry}
        except Exception as e:
            logger.warning("Failed to load user tokens: %s", e)


def get_current_auth_optional(x_admin_token: Optional[str] = None) -> dict:
    """Helper: returns current role and email without raising exception."""
    if not ADMIN_PASSWORD:
        return {"role": "admin", "email": None}
    
    if x_admin_token:
        if _validate_admin_token(x_admin_token):
            return {"role": "admin", "email": None}
        if x_admin_token in user_tokens:
            val = user_tokens[x_admin_token]
            if datetime.now() < val["expiry"]:
                email = val["email"]
                role = "dgbas" if users.get(email, {}).get("role") == "dgbas" else "user"
                return {"role": role, "email": email}
            else:
                del user_tokens[x_admin_token]
                _save_user_tokens()
    return {"role": "guest", "email": None}


def get_current_auth(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
) -> dict:
    """Dependency: require valid admin or user token. Raises 401 if guest."""
    auth = get_current_auth_optional(x_admin_token)
    if auth["role"] == "guest":
        raise HTTPException(
            status_code=401,
            detail="Authentication required",
            headers={"X-Needs-Auth": "true"}
        )
    return auth


def check_admin(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Dependency: require valid admin token. Raises 401 if invalid."""
    if not ADMIN_PASSWORD:
        # No admin password configured → skip auth (local dev mode)
        return True
    if not _validate_admin_token(x_admin_token or ""):
        raise HTTPException(
            status_code=401,
            detail="Admin authentication required",
            headers={"X-Needs-Admin": "true"}
        )
    return True


def check_admin_or_dgbas(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
) -> dict:
    """Dependency: require admin or 主計總處 (dgbas). Returns the auth dict."""
    auth = get_current_auth_optional(x_admin_token)
    if auth["role"] not in ("admin", "dgbas"):
        raise HTTPException(
            status_code=403,
            detail="需要管理員或主計總處權限",
            headers={"X-Needs-Admin": "true"}
        )
    return auth


def _guard_dgbas_target(auth: dict, target_email: str):
    """主計總處 may only manage regular users — never admins or other 主計總處."""
    if auth["role"] == "dgbas" and users.get(target_email, {}).get("role") == "dgbas":
        raise HTTPException(status_code=403, detail="主計總處不可管理其他主計總處帳號")

# In-memory stores
sessions = {}
publish_store = {}      # share_token → session_id
published_forms = {}    # session_id → share_token
response_store = {}     # session_id → [{id, data, submitted_at, respondent}]


class SessionData(BaseModel):
    """Session data for a user's editing session."""
    id: str
    name: str
    created_at: str
    updated_at: str
    original_html: str
    original_json: list
    metadata: dict
    current_data: Optional[list] = None
    form_data: Optional[dict] = None
    email: Optional[str] = ""
    password_hash: Optional[str] = ""
    fund_col: Optional[int] = 0          # which column holds the 基金名稱 (for 彙整)


def _hash_password(pwd: str) -> str:
    """Hash a password with Bcrypt. Returns salted hash string."""
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(pwd.encode('utf-8'), salt).decode('utf-8')


def _verify_password(pwd: str, hashed: str) -> bool:
    """Verify password against a Bcrypt hash, with fallback to SHA-256 for backwards compatibility."""
    if not pwd or not hashed:
        return False
    if hashed.startswith("$2b$") or hashed.startswith("$2a$"):
        try:
            return bcrypt.checkpw(pwd.encode('utf-8'), hashed.encode('utf-8'))
        except Exception:
            return False
    else:
        # Fallback to legacy SHA-256 hash comparison
        legacy_hash = hashlib.sha256(pwd.encode('utf-8')).hexdigest()
        return hmac.compare_digest(legacy_hash, hashed)


def check_session_auth(
    session_id: str,
    password_header: Optional[str] = None,
    password_query: Optional[str] = None,
    x_admin_token: Optional[str] = None,
    write: bool = False
):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    session = sessions[session_id]

    # Bypass verification if the user is admin
    auth = get_current_auth_optional(x_admin_token)
    if auth["role"] == "admin":
        return session
    # 主計總處: read-only access to every project; never allowed to modify
    if auth["role"] == "dgbas":
        if write:
            raise HTTPException(status_code=403, detail="主計總處為唯讀權限，無法修改專案")
        return session

    if session.password_hash:
        pwd = password_header or password_query
        if not pwd:
            raise HTTPException(
                status_code=401,
                detail="Password required",
                headers={"X-Needs-Password": "true"}
            )
        if not _verify_password(pwd, session.password_hash):
            raise HTTPException(status_code=403, detail="Incorrect password")
    return session


class SaveRequest(BaseModel):
    """Request to save edited data."""
    data: list
    name: Optional[str] = None
    fund_col: Optional[int] = None


class TemplateData(BaseModel):
    """Template data for creating a new budget from existing data."""
    name: str
    data: list
    metadata: Optional[dict] = None
    supervisor_name: Optional[str] = ""
    creator_email: Optional[str] = ""


# ── Persistence ──────────────────────────────────────────
PERSIST_FILES = {
    "publish_store": DATA_DIR / "publish_store.json",
    "published_forms": DATA_DIR / "published_forms.json",
    "response_store": DATA_DIR / "response_store.json",
    "sessions_index": DATA_DIR / "sessions_index.json",
}

def _load_persist():
    for key, path in PERSIST_FILES.items():
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding='utf-8'))
                if key == "publish_store":
                    publish_store.update(data)
                elif key == "published_forms":
                    published_forms.update(data)
                elif key == "response_store":
                    response_store.update(data)
                elif key == "sessions_index":
                    for sid in data:
                        sp = DATA_DIR / f"session_{sid}.json"
                        if sp.exists():
                            try:
                                sd = json.loads(sp.read_text(encoding='utf-8'))
                                sessions[sid] = SessionData(**sd)
                            except Exception as e:
                                logger.warning("Failed to load session %s: %s", sid, e)
            except Exception as e:
                logger.warning("Failed to load persist file %s: %s", path, e)

def _save_persist(key):
    d = {"publish_store": publish_store, "published_forms": published_forms,
         "response_store": response_store}[key]
    PERSIST_FILES[key].write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding='utf-8')

def _save_session(sid):
    if sid in sessions:
        (DATA_DIR / f"session_{sid}.json").write_text(
            sessions[sid].model_dump_json(indent=2), encoding='utf-8')
    idx = sorted(sessions.keys())
    PERSIST_FILES["sessions_index"].write_text(
        json.dumps(idx, ensure_ascii=False), encoding='utf-8')

def _cleanup_orphan_uploads():
    """Remove leftover temp files and uploads whose session no longer exists."""
    for f in UPLOAD_DIR.glob("*"):
        if f.name == ".gitkeep" or not f.is_file():
            continue
        # temp_ files are transient parse artifacts; never tied to a live session
        if f.name.startswith("temp_"):
            f.unlink(missing_ok=True)
            continue
        session_prefix = f.name.split("_", 1)[0]
        if session_prefix not in sessions:
            f.unlink(missing_ok=True)


_load_persist()
_load_admin_tokens()
_load_users()
_load_user_tokens()
_cleanup_orphan_uploads()


# ── Authentication & User Management Endpoints ───────────

class AdminLoginRequest(BaseModel):
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


class CreateUserRequest(BaseModel):
    email: str
    password: str
    supervisor_domain: Optional[str] = ""   # registry domain: enterprise | special
    supervisor_code: Optional[str] = ""      # 主管機關編號
    agency_name: Optional[str] = ""          # 機關名稱


@app.post("/api/admin/login")
@limiter.limit("5/minute")
async def admin_login(request: Request, payload: AdminLoginRequest):
    """Authenticate as admin. Returns a session token (Backward compatibility)."""
    if not ADMIN_PASSWORD:
        return {"success": True, "token": "", "message": "No admin password configured"}
    if payload.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="管理員密碼錯誤")
    token = _generate_admin_token()
    return {
        "success": True,
        "token": token,
        "expires_in_hours": ADMIN_TOKEN_EXPIRY_HOURS
    }


@app.get("/api/admin/verify")
async def admin_verify(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Check if current admin token is valid (Backward compatibility)."""
    if not ADMIN_PASSWORD:
        return {"authenticated": True, "admin_required": False}
    valid = _validate_admin_token(x_admin_token or "")
    return {"authenticated": valid, "admin_required": True}


@app.post("/api/admin/logout")
async def admin_logout(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Invalidate an admin token (Backward compatibility)."""
    if x_admin_token and x_admin_token in admin_tokens:
        del admin_tokens[x_admin_token]
        _save_admin_tokens()
    return {"success": True}


@app.post("/api/auth/login")
@limiter.limit("5/minute")
async def auth_login(request: Request, payload: LoginRequest):
    """Unified login for admin and users."""
    email = payload.email.strip()
    password = payload.password
    
    is_admin = False
    if ADMIN_PASSWORD and password == ADMIN_PASSWORD:
        if not email or email.lower() == "admin":
            is_admin = True
            
    if not ADMIN_PASSWORD:
        is_admin = True
        
    if is_admin:
        token = _generate_admin_token()
        return {
            "success": True,
            "role": "admin",
            "token": token,
            "expires_in_hours": ADMIN_TOKEN_EXPIRY_HOURS
        }
        
    # Check regular user
    email_lower = email.lower()
    if email_lower in users:
        user_info = users[email_lower]
        if _verify_password(password, user_info["password_hash"]):
            token = secrets.token_hex(24)
            expiry = datetime.now() + timedelta(hours=ADMIN_TOKEN_EXPIRY_HOURS)
            user_tokens[token] = {"email": email_lower, "expiry": expiry}
            _save_user_tokens()
            role = "dgbas" if user_info.get("role") == "dgbas" else "user"
            return {
                "success": True,
                "role": role,
                "email": email_lower,
                "token": token,
                "expires_in_hours": ADMIN_TOKEN_EXPIRY_HOURS
            }
            
    raise HTTPException(status_code=403, detail="帳號或密碼錯誤")


@app.get("/api/auth/verify")
async def auth_verify(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Check token status and return role + email."""
    if not ADMIN_PASSWORD:
        return {"authenticated": True, "role": "admin", "admin_required": False}
        
    auth = get_current_auth_optional(x_admin_token)
    if auth["role"] == "guest":
        return {"authenticated": False, "role": "guest", "admin_required": True}
        
    return {
        "authenticated": True,
        "role": auth["role"],
        "email": auth.get("email"),
        "admin_required": True
    }


@app.post("/api/auth/logout")
async def auth_logout(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Invalidate any token (admin or user)."""
    if x_admin_token:
        if x_admin_token in admin_tokens:
            del admin_tokens[x_admin_token]
            _save_admin_tokens()
        elif x_admin_token in user_tokens:
            del user_tokens[x_admin_token]
            _save_user_tokens()
    return {"success": True}


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@app.put("/api/auth/change-password")
async def change_password(
    payload: ChangePasswordRequest,
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Let an authenticated user change their own password."""
    auth = get_current_auth_optional(x_admin_token)
    if auth["role"] not in ("user", "dgbas") or not auth["email"]:
        raise HTTPException(status_code=403, detail="Only regular users can change password here")
    email = auth["email"]
    if email not in users:
        raise HTTPException(status_code=404, detail="User not found")
    if not _verify_password(payload.old_password, users[email]["password_hash"]):
        raise HTTPException(status_code=403, detail="舊密碼錯誤")
    if len(payload.new_password) < 1:
        raise HTTPException(status_code=400, detail="新密碼不可為空")
    users[email]["password_hash"] = _hash_password(payload.new_password)
    _save_users()
    return {"success": True}


# User maintenance endpoints (Admin or 主計總處)
@app.get("/api/admin/users")
async def list_users(auth=Depends(check_admin_or_dgbas)):
    """List all registered users. Admin or 主計總處."""
    return {
        "users": [
            {
                "email": email,
                "created_at": info.get("created_at", ""),
                "supervisor": info.get("supervisor"),
                "agency_name": info.get("agency_name", ""),
                "role": info.get("role", "")
            }
            for email, info in users.items()
        ]
    }


@app.post("/api/admin/users")
async def create_user(payload: CreateUserRequest, auth=Depends(check_admin_or_dgbas)):
    """Create a new user (always a regular publishing user). Admin or 主計總處."""
    email = payload.email.strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email cannot be empty")
    
    if "@" not in email:
        raise HTTPException(status_code=400, detail="Invalid email address")
        
    if email in users:
        raise HTTPException(status_code=400, detail="User already exists")
        
    pwd_hash = _hash_password(payload.password)
    users[email] = {
        "password_hash": pwd_hash,
        "created_at": datetime.now().isoformat(),
        "supervisor": _resolve_supervisor(payload.supervisor_domain, payload.supervisor_code),
        "agency_name": (payload.agency_name or "").strip()
    }
    _save_users()
    return {"success": True, "message": f"User {email} created successfully"}


@app.put("/api/admin/users/{email}/profile")
async def set_user_profile(email: str, payload: CreateUserRequest, auth=Depends(check_admin_or_dgbas)):
    """Update a user's 主管機關 and 機關名稱. Admin or 主計總處."""
    email = email.strip().lower()
    if email not in users:
        raise HTTPException(status_code=404, detail="User not found")
    _guard_dgbas_target(auth, email)
    users[email]["supervisor"] = _resolve_supervisor(payload.supervisor_domain, payload.supervisor_code)
    users[email]["agency_name"] = (payload.agency_name or "").strip()
    _save_users()
    return {"success": True, "supervisor": users[email]["supervisor"], "agency_name": users[email]["agency_name"]}


@app.delete("/api/admin/users/{email}")
async def delete_user(email: str, auth=Depends(check_admin_or_dgbas)):
    """Delete a user. Admin or 主計總處 (主計總處 cannot delete other 主計總處)."""
    email = email.strip().lower()
    if email not in users:
        raise HTTPException(status_code=404, detail="User not found")
    _guard_dgbas_target(auth, email)

    del users[email]
    _save_users()
    
    # Revoke tokens for deleted user
    tokens_to_del = [t for t, val in user_tokens.items() if val["email"] == email]
    for t in tokens_to_del:
        del user_tokens[t]
    if tokens_to_del:
        _save_user_tokens()

    return {"success": True, "message": f"User {email} deleted successfully"}


class SetRoleRequest(BaseModel):
    role: str = ""   # "dgbas" (主計總處) or "" (regular user)


@app.put("/api/admin/users/{email}/role")
async def set_user_role(email: str, payload: SetRoleRequest, _admin=Depends(check_admin)):
    """Grant or revoke the 主計總處 (dgbas) role. Admin only — guards against escalation."""
    email = email.strip().lower()
    if email not in users:
        raise HTTPException(status_code=404, detail="User not found")
    role = (payload.role or "").strip()
    if role not in ("", "dgbas"):
        raise HTTPException(status_code=400, detail="Invalid role")
    if role:
        users[email]["role"] = role
    else:
        users[email].pop("role", None)
    _save_users()
    # Revoke existing tokens so the new role takes effect on next login
    tokens_to_del = [t for t, val in user_tokens.items() if val["email"] == email]
    for t in tokens_to_del:
        del user_tokens[t]
    if tokens_to_del:
        _save_user_tokens()
    return {"success": True, "role": role}


@app.get("/api/admin/audit-log")
async def get_audit_log(
    limit: int = 500,
    email: str = "",
    action: str = "",
    auth=Depends(check_admin_or_dgbas)
):
    """Return recent audit-log entries (newest first). Admin or 主計總處."""
    limit = max(1, min(limit, 2000))
    return {"entries": _read_audit(limit, email, action)}


# ── Fund / supervising-authority registry ────────────────
_active_funds_cache = None

def _load_active_funds():
    global _active_funds_cache
    if _active_funds_cache is None:
        p = Path(__file__).parent / "active_funds.json"
        if p.exists():
            _active_funds_cache = json.loads(p.read_text(encoding="utf-8"))
        else:
            _active_funds_cache = {}
    return _active_funds_cache


def _resolve_supervisor(domain, code):
    """Build a supervisor record {domain, code, name} from the registry, or None."""
    domain = (domain or "").strip()
    code = (code or "").strip()
    if not domain or not code:
        return None
    dd = registry.get_registry().get("domains", {}).get(domain, {})
    name = dd.get("supervisors", {}).get(code)
    if not name:
        return None
    return {"domain": domain, "code": code, "name": name}


@app.get("/api/registry")
async def get_registry_api():
    """Return the full fund/supervisor registry (official names, codes, aliases)."""
    return registry.get_registry()


def _flatten_fund_names(af):
    """Flatten active funds (both domains use {name, children} objects) to a set of names."""
    names = set()
    for dom in ("enterprise", "special"):
        for entry in af.get(dom, {}).get("funds", []):
            if isinstance(entry, str):
                names.add(entry)
            elif isinstance(entry, dict):
                names.add(entry["name"])
                for child in entry.get("children", []):
                    names.add(child)
    return names


@app.get("/api/fund-names")
async def fund_names_api():
    """Return a flat list of active fund names + aliases for autocomplete."""
    af = _load_active_funds()
    names = _flatten_fund_names(af)
    for alias in af.get("aliasMap", {}):
        names.add(alias)
    return {"names": sorted(names)}


@app.get("/api/supervisors")
async def list_supervisors_api():
    """Flat list of supervising authorities (主管機關) across both domains."""
    return {"supervisors": registry.list_supervisors()}


class SupervisorPayload(BaseModel):
    domain: str
    code: str
    name: str


@app.put("/api/admin/supervisors")
async def upsert_supervisor_api(payload: SupervisorPayload, _admin=Depends(check_admin)):
    try:
        registry.upsert_supervisor(payload.domain, payload.code.strip(), payload.name.strip())
    except KeyError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"success": True}


@app.delete("/api/admin/supervisors/{domain}/{code}")
async def delete_supervisor_api(domain: str, code: str, _admin=Depends(check_admin)):
    try:
        registry.delete_supervisor(domain, code)
    except KeyError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"success": True}


@app.get("/api/consolidate")
async def consolidate(_admin=Depends(check_admin)):
    """Cross-project consolidation grouped 主管機關 → 機關 → 基金. Admin only.

    主管機關 / 機關名稱 come from the project owner's user profile; 基金 is taken
    from each row's designated fund column and normalized against the official
    registry (alias + fuzzy match) to obtain the canonical name and 基金編號.
    """
    UNSET = ("zzz", "（未設定主管機關）")
    groups = {}   # (sup_code, sup_name) -> { agency -> { (fund_code, fund_name) -> node } }

    for sid, session in sessions.items():
        responses = response_store.get(sid, [])
        if not responses:
            continue
        owner = (session.email or "").strip().lower()
        uinfo = users.get(owner, {})
        sup = uinfo.get("supervisor") or {}
        sup_code = sup.get("code") or ""
        sup_key = (sup_code, sup.get("name") or "") if sup_code else UNSET
        agency = uinfo.get("agency_name") or session.email or "（未指定機關）"
        fund_col = session.fund_col or 0

        for r in responses:
            rdata = r.get("data", []) or []
            header = [c.get("value", "") for c in (rdata[0] if rdata else [])]
            for ri in range(1, len(rdata)):
                row = rdata[ri] or []
                cell = row[fund_col] if fund_col < len(row) else None
                fund_raw = (cell.get("value") or "").strip() if isinstance(cell, dict) else ""
                if not fund_raw:
                    continue
                m = registry.match_fund(fund_raw)
                fund_code = m["code"] if m else ""
                fund_name = m["name"] if m else fund_raw
                entry = {
                    "session_id": sid, "session_name": session.name,
                    "respondent": r.get("respondent", ""), "email": r.get("email", ""),
                    "submitted_at": r.get("submitted_at", ""),
                    "header": header,
                    "values": [c.get("value", "") if isinstance(c, dict) else "" for c in row],
                    "fund_raw": fund_raw,
                }
                fk = (fund_code or "zzz", fund_name)
                node = groups.setdefault(sup_key, {}).setdefault(agency, {}).setdefault(
                    fk, {"fund_code": fund_code, "fund_name": fund_name,
                         "matched": bool(m), "rows": []})
                node["rows"].append(entry)

    out_groups = []
    for sup_key in sorted(groups.keys(), key=lambda k: k[0]):
        agencies = []
        for agency in sorted(groups[sup_key].keys()):
            funds = [groups[sup_key][agency][fk] for fk in sorted(groups[sup_key][agency].keys())]
            agencies.append({"agency": agency, "funds": funds})
        out_groups.append({
            "supervisor_code": "" if sup_key[0] == "zzz" else sup_key[0],
            "supervisor_name": sup_key[1],
            "agencies": agencies,
        })
    total_rows = sum(len(f["rows"]) for g in out_groups for a in g["agencies"] for f in a["funds"])
    return {"groups": out_groups, "total_rows": total_rows}


@app.get("/api/active-funds")
async def get_active_funds_api(_admin=Depends(check_admin)):
    """Return the active funds config. Admin only."""
    return _load_active_funds()


@app.post("/api/active-funds")
async def save_active_funds_api(request: Request, _admin=Depends(check_admin)):
    """Save updated active funds config. Admin only."""
    global _active_funds_cache
    payload = await request.json()
    p = Path(__file__).parent / "active_funds.json"
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _active_funds_cache = payload
    return {"success": True}


@app.get("/api/fund-coverage")
async def fund_coverage_api(_admin=Depends(check_admin)):
    """Compare uploaded response fund names against active funds list. Admin only."""
    af = _load_active_funds()
    alias_map = af.get("aliasMap", {})

    uploaded = set()
    for sid, session in sessions.items():
        responses = response_store.get(sid, [])
        fund_col = session.fund_col or 0
        for r in responses:
            rdata = r.get("data", []) or []
            for ri in range(1, len(rdata)):
                row = rdata[ri] or []
                cell = row[fund_col] if fund_col < len(row) else None
                raw = (cell.get("value") or "").strip() if isinstance(cell, dict) else ""
                if not raw:
                    continue
                canonical = alias_map.get(raw)
                if canonical:
                    uploaded.add(canonical)
                m = registry.match_fund(raw)
                uploaded.add(m["name"] if m else raw)

    result = {}
    for dom in ("enterprise", "special"):
        dd = af.get(dom, {})
        dom_funds = []
        total = 0
        covered = 0
        for entry in dd.get("funds", []):
            if isinstance(entry, str):
                dom_funds.append({"name": entry, "uploaded": entry in uploaded})
                total += 1
                if entry in uploaded:
                    covered += 1
            elif isinstance(entry, dict):
                parent_name = entry["name"]
                children = entry.get("children", [])
                kids = [{"name": c, "uploaded": c in uploaded} for c in children]
                dom_funds.append({"name": parent_name, "uploaded": parent_name in uploaded, "children": kids})
                total += 1 + len(kids)
                if parent_name in uploaded:
                    covered += 1
                covered += sum(1 for k in kids if k["uploaded"])
        result[dom] = {
            "label": dd.get("label", dom),
            "total": total,
            "covered": covered,
            "funds": dom_funds,
        }
    return result


@app.get("/")
async def root():
    """Serve the main frontend page (project management)."""
    frontend_path = Path(__file__).parent.parent / "frontend" / "index.html"
    if frontend_path.exists():
        html = frontend_path.read_text(encoding='utf-8')
        admin_required = 'true' if ADMIN_PASSWORD else 'false'
        head_script = f'<script>window.LANDING_MODE="projects";window.ADMIN_REQUIRED={admin_required};</script>'
        head_style = '<style>#editor{display:none!important}#fill-mode{display:none!important}#fill-banner{display:none!important}</style>'
        html = html.replace("</head>", f"{head_script}{head_style}</head>")
        return HTMLResponse(content=html, headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"})
    return {"message": "Budget Table Editor API", "version": "1.0.0"}


@app.get("/create")
async def create_page():
    """Serve project creation page without listing other projects."""
    frontend_path = Path(__file__).parent.parent / "frontend" / "index.html"
    if frontend_path.exists():
        html = frontend_path.read_text(encoding='utf-8')
        head_script = '<script>window.LANDING_MODE="create";window.ADMIN_REQUIRED=false;</script>'
        head_style = '<style>#editor{display:none!important}#fill-mode{display:none!important}#fill-banner{display:none!important}</style>'
        html = html.replace("</head>", f"{head_script}{head_style}</head>")
        return HTMLResponse(content=html, headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"})
    return {"message": "Budget Table Editor API", "version": "1.0.0"}


@app.get("/editor/{session_id}")
async def editor_page(session_id: str):
    """Serve editor page for a specific session."""
    frontend_path = Path(__file__).parent.parent / "frontend" / "index.html"
    if not frontend_path.exists():
        raise HTTPException(status_code=404, detail="Page not found")
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    html = frontend_path.read_text(encoding='utf-8')
    # Inject EDIT_SESSION_ID and hide projects/fill, show editor
    admin_required = 'true' if ADMIN_PASSWORD else 'false'
    head_script = f'<script>window.EDIT_SESSION_ID="{session_id}";window.ADMIN_REQUIRED={admin_required};</script>'
    head_style = '<style>#projects-section{display:none!important}#fill-mode{display:none!important}#fill-banner{display:none!important}#editor{display:block!important}</style>'
    html = html.replace("</head>", f"{head_script}{head_style}</head>")
    return HTMLResponse(content=html, headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"})


@app.post("/api/upload-xlsx")
@limiter.limit("10/minute")
async def upload_xlsx(
    request: Request,
    file: UploadFile = File(...),
    sheet_index: int = Form(0),
    mode: str = Form("table"),  # 'table' or 'form'
    email: str = Form(""),
    password: str = Form(""),
    name: str = Form(""),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Upload an XLSX file and convert to HTML table or form (Authenticated)."""
    auth = get_current_auth(x_admin_token)
    if auth["role"] == "user":
        email = auth["email"]
    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    # Check file size
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large. Maximum {MAX_UPLOAD_BYTES // (1024*1024)}MB allowed.")
    await file.seek(0)

    # Save uploaded file
    session_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{session_id}_{file.filename}"

    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Process the XLSX file
        result = process_xlsx(str(file_path), sheet_index, mode=mode)

        password_hash = ""
        if password:
            password_hash = _hash_password(password)

        if mode == 'form':
            # Form mode
            session = SessionData(
                id=session_id,
                name=name if name else file.filename,
                created_at=datetime.now().isoformat(),
                updated_at=datetime.now().isoformat(),
                original_html="",
                original_json=[],
                metadata=result['metadata'],
                current_data=[],
                form_data=result.get('form'),
                email=email,
                password_hash=password_hash
            )

            sessions[session_id] = session
            _save_session(session_id)

            return {
                "success": True,
                "session_id": session_id,
                "form": result.get('form'),
                "metadata": result['metadata']
            }
        else:
            # Table mode (default)
            session = SessionData(
                id=session_id,
                name=name if name else file.filename,
                created_at=datetime.now().isoformat(),
                updated_at=datetime.now().isoformat(),
                original_html=result['html'],
                original_json=result['json'],
                metadata=result['metadata'],
                current_data=result['json'],
                email=email,
                password_hash=password_hash
            )

            sessions[session_id] = session
            _save_session(session_id)

            return {
                "success": True,
                "session_id": session_id,
                "html": result['html'],
                "json": result['json'],
                "metadata": result['metadata']
            }

    except Exception as e:
        # Clean up file on error
        if file_path.exists():
            file_path.unlink()
        logger.error("Error processing uploaded file (session %s): %s", session_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail="無法處理此檔案，請確認為有效的 XLSX 檔案")


@app.get("/api/sessions")
async def list_sessions(auth=Depends(get_current_auth)):
    """List active editing sessions. Admin/主計總處 see all, users see only their own."""
    if auth["role"] in ("admin", "dgbas"):
        visible_sessions = list(sessions.values())
    elif auth["role"] == "user":
        visible_sessions = [s for s in sessions.values() if s.email == auth["email"]]
    else:
        visible_sessions = []

    return {
        "sessions": [
            {
                "id": s.id,
                "name": s.name,
                "created_at": s.created_at,
                "updated_at": s.updated_at,
                "row_count": s.metadata.get('row_count', 0),
                "col_count": s.metadata.get('col_count', 0),
                "published": s.id in published_forms,
                "share_token": published_forms.get(s.id),
                "response_count": len(response_store.get(s.id, []))
            }
            for s in visible_sessions
        ]
    }


@app.get("/api/sessions/{session_id}")
@limiter.limit("10/minute")
async def get_session(
    request: Request,
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Get session data by ID. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token)
    return {
        "id": session.id,
        "name": session.name,
        "html": session.original_html,
        "json": session.original_json,
        "current_data": session.current_data or session.original_json,
        "metadata": session.metadata,
        "form": session.form_data,
        "fund_col": session.fund_col or 0,
        "created_at": session.created_at,
        "updated_at": session.updated_at
    }


@app.post("/api/sessions/{session_id}/save")
async def save_session(
    session_id: str,
    request: SaveRequest,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Save edited data to session. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)
    session.current_data = request.data
    if request.name:
        session.name = request.name
    if request.fund_col is not None:
        session.fund_col = request.fund_col
    session.updated_at = datetime.now().isoformat()
    _save_session(session_id)

    return {"success": True, "updated_at": session.updated_at}


@app.get("/api/sessions/{session_id}/export/json")
async def export_json(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Export session data as JSON file. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token)
    data = {
        "session": {
            "id": session.id,
            "name": session.name,
            "created_at": session.created_at,
            "updated_at": session.updated_at
        },
        "metadata": session.metadata,
        "data": session.current_data or session.original_json
    }

    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    filename = f"budget_export_{session.name.replace('.xlsx', '')}_{session_id}.json"
    # RFC 5987 encoding so non-ASCII (e.g. 中文) filenames survive in Content-Disposition
    disposition = f"attachment; filename=\"export_{session_id}.json\"; filename*=UTF-8''{quote(filename)}"
    return Response(
        content=json_str.encode('utf-8'),
        media_type="application/json",
        headers={"Content-Disposition": disposition}
    )


@app.post("/api/sessions/{session_id}/import/json")
async def import_json(
    session_id: str,
    file: UploadFile = File(...),
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Import JSON data into session."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)
    try:
        content = await file.read()
        data = json.loads(content.decode('utf-8'))

        if 'data' in data:
            session.current_data = data['data']
        session.updated_at = datetime.now().isoformat()
        _save_session(session_id)

        return {"success": True, "message": "JSON imported successfully"}

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON file")
    except Exception as e:
        logger.error("Error importing JSON (session %s): %s", session_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail="匯入 JSON 失敗")


@app.post("/api/sessions/{session_id}/reset")
async def reset_session(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Reset session to original data. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)
    session.current_data = session.original_json

    # Reset form_data to original by re-processing the uploaded file
    if session.form_data:
        pattern = str(UPLOAD_DIR / f"{session_id}_*.xlsx")
        files = glob.glob(pattern)
        if files:
            try:
                result = process_xlsx(files[0], mode='form')
                session.form_data = result.get('form')
            except Exception as e:
                logger.warning("Failed to re-process XLSX for reset (session %s): %s", session_id, e)

    session.updated_at = datetime.now().isoformat()
    _save_session(session_id)
    return {"success": True, "message": "Session reset to original data"}


@app.delete("/api/sessions/{session_id}")
async def delete_session(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Delete a session and all associated data. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)
    del sessions[session_id]

    # Clean up session JSON file
    session_file = DATA_DIR / f"session_{session_id}.json"
    session_file.unlink(missing_ok=True)

    # Clean up publish/response stores
    if session_id in published_forms:
        token = published_forms.pop(session_id)
        publish_store.pop(token, None)
        _save_persist("publish_store")
        _save_persist("published_forms")

    if session_id in response_store:
        del response_store[session_id]
        _save_persist("response_store")

    # Clean up uploaded XLSX files
    for f in UPLOAD_DIR.glob(f"{session_id}_*"):
        f.unlink(missing_ok=True)

    _save_session(session_id)  # updates index
    return {"success": True, "message": "Session deleted"}


@app.post("/api/templates/save")
async def save_template(request: TemplateData, _admin=Depends(check_admin)):
    """Save current data as a reusable template. Requires admin auth."""
    template_id = str(uuid.uuid4())[:8]
    template_path = TEMPLATE_DIR / f"template_{template_id}.json"

    template_data = {
        "id": template_id,
        "name": request.name,
        "created_at": datetime.now().isoformat(),
        "data": request.data,
        "metadata": request.metadata or {},
        "supervisor_name": request.supervisor_name or "",
        "creator_email": request.creator_email or "",
    }

    template_path.write_text(json.dumps(template_data, ensure_ascii=False, indent=2), encoding='utf-8')

    return {"success": True, "template_id": template_id}


@app.post("/api/templates/save-from-session/{session_id}")
async def save_template_from_session(session_id: str, _admin=Depends(check_admin)):
    """Save an existing project as a reusable template."""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    s = sessions[session_id]
    email = s.email or ""
    supervisor_name = ""
    if email and email in users:
        sup = users[email].get("supervisor")
        if sup:
            supervisor_name = sup.get("name", "")

    template_id = str(uuid.uuid4())[:8]
    template_path = TEMPLATE_DIR / f"template_{template_id}.json"
    tpl = {
        "id": template_id,
        "name": s.name,
        "created_at": datetime.now().isoformat(),
        "data": s.current_data or s.original_json,
        "metadata": s.metadata or {},
        "supervisor_name": supervisor_name,
        "creator_email": email,
    }
    template_path.write_text(json.dumps(tpl, ensure_ascii=False, indent=2), encoding='utf-8')
    return {"success": True, "template_id": template_id, "name": s.name}


@app.get("/api/templates")
async def list_templates(auth=Depends(check_admin_or_dgbas)):
    """List all saved templates. Admin or 主計總處."""
    templates = []
    for f in TEMPLATE_DIR.glob("template_*.json"):
        try:
            data = json.loads(f.read_text(encoding='utf-8'))
            templates.append({
                "id": data.get('id'),
                "name": data.get('name'),
                "created_at": data.get('created_at'),
                "row_count": len(data.get('data', [])),
                "supervisor_name": data.get('supervisor_name', ''),
                "creator_email": data.get('creator_email', ''),
            })
        except Exception as e:
            logger.warning("Failed to load template %s: %s", f.name, e)
            continue

    return {"templates": templates}


@app.get("/api/templates/{template_id}")
async def get_template(template_id: str, auth=Depends(check_admin_or_dgbas)):
    """Read a single template's full content (read-only). Admin or 主計總處."""
    template_path = TEMPLATE_DIR / f"template_{template_id}.json"
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="Template not found")
    return json.loads(template_path.read_text(encoding='utf-8'))


@app.delete("/api/templates/{template_id}")
async def delete_template(template_id: str, _admin=Depends(check_admin)):
    """Delete a template. Requires admin auth."""
    template_path = TEMPLATE_DIR / f"template_{template_id}.json"
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="Template not found")
    template_path.unlink()
    return {"success": True}


@app.post("/api/templates/{template_id}/load")
async def load_template(template_id: str, _admin=Depends(check_admin)):
    """Load a template and create a new session from it. Requires admin auth."""
    template_path = TEMPLATE_DIR / f"template_{template_id}.json"
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="Template not found")

    template_data = json.loads(template_path.read_text(encoding='utf-8'))

    session_id = str(uuid.uuid4())[:8]
    session = SessionData(
        id=session_id,
        name=template_data.get('name', 'Untitled'),
        created_at=datetime.now().isoformat(),
        updated_at=datetime.now().isoformat(),
        original_html="",
        original_json=template_data.get('data', []),
        metadata=template_data.get('metadata', {}),
        current_data=template_data.get('data', [])
    )

    sessions[session_id] = session

    _save_session(session_id)
    return {"success": True, "session_id": session_id}


# ========== Publish / Fill / Responses ==========

class PublishResponse(BaseModel):
    share_token: str
    fill_url: str
    published_at: str
    response_count: int

class SubmitRequest(BaseModel):
    data: list
    respondent: Optional[str] = ""
    email: Optional[str] = ""
    password: Optional[str] = ""


class ExportXlsxRequest(BaseModel):
    name: str
    data: list


class LoadResponseRequest(BaseModel):
    email: str
    password: str

@app.post("/api/sessions/{session_id}/publish")
async def publish_form(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Publish a form for others to fill. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)

    if session_id in published_forms:
        token = published_forms[session_id]
    else:
        # Unguessable share token (≈128-bit) — this token is the only access
        # control for public fill links, so it must resist enumeration.
        token = secrets.token_urlsafe(16)
        publish_store[token] = session_id
        published_forms[session_id] = token
        _save_persist("publish_store")
        _save_persist("published_forms")

    return PublishResponse(
        share_token=token,
        fill_url=f"/fill/{token}",
        published_at=datetime.now().isoformat(),
        response_count=len(response_store.get(session_id, []))
    )

@app.get("/api/sessions/{session_id}/publish")
async def get_publish_status(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Get publish status for a session. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token)

    if session_id in published_forms:
        token = published_forms[session_id]
        return PublishResponse(
            share_token=token,
            fill_url=f"/fill/{token}",
            published_at=datetime.now().isoformat(),
            response_count=len(response_store.get(session_id, []))
        )
    return {"published": False}

@app.delete("/api/sessions/{session_id}/publish")
async def unpublish_form(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Unpublish a form. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)
    if session_id in published_forms:
        token = published_forms.pop(session_id)
        publish_store.pop(token, None)
        _save_persist("published_forms")
        _save_persist("publish_store")
    return {"success": True, "published": False}

@app.get("/fill/{token}")
async def fill_form_page(token: str):
    """Serve fill-mode HTML page."""
    frontend_path = Path(__file__).parent.parent / "frontend" / "index.html"
    if not frontend_path.exists():
        raise HTTPException(status_code=404, detail="Page not found")
    html = frontend_path.read_text(encoding='utf-8')
    # Inject fill token and CSS to hide other sections, show fill mode
    head_script = f'<script>window.FILL_TOKEN="{token}";</script>'
    head_style = '<style>#projects-section{display:none!important}#editor{display:none!important}#fill-mode{display:block!important}#fill-banner{display:block!important}</style>'
    html = html.replace("</head>", f"{head_script}{head_style}</head>")
    return HTMLResponse(content=html, headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"})

@app.get("/api/fill/{token}/data")
@limiter.limit("30/minute")
async def get_fill_data(token: str, request: Request):
    """Get form data for fill mode."""
    if token not in publish_store:
        raise HTTPException(status_code=404, detail="Form not found or has been unpublished")
    session_id = publish_store[token]
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    session = sessions[session_id]
    data = session.current_data or session.original_json
    return {
        "session_id": session_id,
        "name": session.name,
        "data": data,
        "metadata": session.metadata,
        "fund_col": session.fund_col if session.fund_col is not None else 0
    }

@app.post("/api/fill/{token}/submit")
@limiter.limit("10/minute")
async def submit_fill(token: str, request: Request, payload: SubmitRequest):
    """Submit a filled form response."""
    if token not in publish_store:
        raise HTTPException(status_code=404, detail="Form not found or has been unpublished")
    session_id = publish_store[token]

    respondent_email = payload.email.strip() if payload.email else ""
    password_hash = ""
    if payload.password:
        password_hash = _hash_password(payload.password)

    # 尋找是否已有同一個 Email 的填寫記錄 (自動覆蓋修正)
    existing_resp = None
    if session_id in response_store and respondent_email:
        for r in response_store[session_id]:
            if r.get("email") == respondent_email:
                existing_resp = r
                break

    if existing_resp:
        # 如果已經填過，必須校驗密碼
        if not payload.password or not _verify_password(payload.password, existing_resp.get("password_hash", "")):
            raise HTTPException(
                status_code=403,
                detail="此 Email 已有填表紀錄。如欲修改，請輸入您當初設定的填表密碼。"
            )

        # 密碼正確，直接覆蓋修正該項目
        existing_resp["data"] = payload.data
        existing_resp["respondent"] = payload.respondent or ""
        existing_resp["modified"] = True
        existing_resp["modified_at"] = datetime.now().isoformat()
        _save_persist("response_store")

        return {
            "success": True,
            "response_id": existing_resp["id"],
            "message": "修正成功",
            "modified": True,
            "submitted_at": existing_resp["modified_at"]
        }

    # 全新提交
    resp = {
        "id": str(uuid.uuid4())[:8],
        "session_id": session_id,
        "data": payload.data,
        "respondent": payload.respondent or "",
        "email": respondent_email,
        "password_hash": password_hash,
        "submitted_at": datetime.now().isoformat(),
        "modified": False,
        "modified_at": None
    }
    if session_id not in response_store:
        response_store[session_id] = []
    response_store[session_id].append(resp)
    _save_persist("response_store")

    return {
        "success": True,
        "response_id": resp["id"],
        "modified": False,
        "submitted_at": resp["submitted_at"]
    }


@app.post("/api/fill/{token}/load-response")
@limiter.limit("10/minute")
async def load_filler_response(token: str, request: Request, payload: LoadResponseRequest):
    """Load a previous submission for a filler to modify."""
    if token not in publish_store:
        raise HTTPException(status_code=404, detail="Form not found")
    session_id = publish_store[token]

    responses = response_store.get(session_id, [])
    for r in responses:
        if r.get("email") == payload.email.strip():
            if _verify_password(payload.password, r.get("password_hash", "")):
                return {
                    "success": True,
                    "respondent": r.get("respondent", ""),
                    "data": r.get("data", [])
                }
            else:
                raise HTTPException(status_code=403, detail="密碼錯誤，無法載入資料")

    raise HTTPException(status_code=404, detail="找不到此 Email 的填表紀錄")

@app.get("/api/sessions/{session_id}/responses")
async def list_responses(
    session_id: str,
    full: bool = False,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """List all submitted responses for a session. Protected by project password or auth.
    
    Args:
        full: If true, include full response data (avoids N+1 queries).
    """
    session = check_session_auth(session_id, x_project_password, None, x_admin_token)
    responses = response_store.get(session_id, [])
    
    if full:
        return {
            "session_id": session_id,
            "count": len(responses),
            "responses": responses
        }
    
    return {
        "session_id": session_id,
        "count": len(responses),
        "responses": [
            {
                "id": r["id"],
                "submitted_at": r["submitted_at"],
                "respondent": r["respondent"],
                "email": r.get("email", ""),
                "modified": r.get("modified", False),
                "modified_at": r.get("modified_at")
            }
            for r in responses
        ]
    }

@app.get("/api/sessions/{session_id}/responses/{response_id}")
async def get_response(
    session_id: str,
    response_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Get a single response with full data. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token)
    responses = response_store.get(session_id, [])
    for r in responses:
        if r["id"] == response_id:
            return r
    raise HTTPException(status_code=404, detail="Response not found")

@app.delete("/api/sessions/{session_id}/responses/{response_id}")
async def delete_response(
    session_id: str,
    response_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Delete a single response. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token, write=True)
    responses = response_store.get(session_id, [])
    for i, r in enumerate(responses):
        if r["id"] == response_id:
            del response_store[session_id][i]
            _save_persist("response_store")
            return {"success": True, "deleted": response_id}
    raise HTTPException(status_code=404, detail="Response not found")

@app.get("/api/sessions/{session_id}/responses/export/csv")
@app.post("/api/sessions/{session_id}/responses/export/csv")
async def export_responses_csv(
    session_id: str,
    x_project_password: Optional[str] = Header(None, alias="X-Project-Password"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Export all responses as CSV. Protected by project password or auth."""
    session = check_session_auth(session_id, x_project_password, None, x_admin_token)
    responses = response_store.get(session_id, [])
    if not responses:
        raise HTTPException(status_code=404, detail="No responses to export")

    # Build CSV: headers from original data row 0, then each response
    sess_data = session.current_data or session.original_json
    headers = [c.get("value", "") for c in (sess_data[0] if sess_data else [])]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["回應時間", "填表人"] + headers)

    for r in responses:
        row_data = r.get("data", [])
        # 排除 row_data[0]（標題列），匯出所有填寫的數據列
        for ri in range(1, len(row_data)):
            row = row_data[ri]
            vals = [c.get("value", "") for c in row]
            writer.writerow([r["submitted_at"], r["respondent"]] + vals)

    csv_content = output.getvalue()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=responses_{session_id}.csv"}
    )


@app.post("/api/parse-xlsx")
@limiter.limit("10/minute")
async def parse_xlsx_only(
    request: Request,
    file: UploadFile = File(...),
    sheet_index: int = Form(0),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token")
):
    """Parse an XLSX file and return raw JSON data without saving any session (Authenticated)."""
    get_current_auth(x_admin_token)
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    temp_id = str(uuid.uuid4())[:8]
    temp_path = UPLOAD_DIR / f"temp_{temp_id}_{file.filename}"
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        result = process_xlsx(str(temp_path), sheet_index, mode='table')
        return {
            "success": True,
            "json": result['json'],
            "metadata": result['metadata']
        }
    except Exception as e:
        logger.error("Error parsing XLSX (%s): %s", temp_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail="無法解析此 XLSX 檔案")
    finally:
        if temp_path.exists():
            temp_path.unlink()


@app.post("/api/export-xlsx")
@limiter.limit("15/minute")
async def export_xlsx_api(
    request: Request,
    payload: ExportXlsxRequest
):
    """Generate and return an XLSX file from the budget table JSON data."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Limit sheet title length to 30 characters (Excel limit is 31)
        sheet_title = payload.name.replace(".xlsx", "").replace(".xls", "")[:30]
        # Excel sheet name cannot contain chars like: \ / ? * : [ ]
        for char in ['\\', '/', '?', '*', ':', '[', ']']:
            sheet_title = sheet_title.replace(char, "_")
        if not sheet_title.strip():
            sheet_title = "Sheet1"
            
        ws = wb.create_sheet(title=sheet_title)
        
        # First write values and styles
        for row_idx, row in enumerate(payload.data):
            if not isinstance(row, list):
                continue
            for col_idx, c in enumerate(row):
                if not isinstance(c, dict):
                    continue
                r_idx = c.get('row') if c.get('row') is not None else (row_idx + 1)
                c_idx = c.get('col') if c.get('col') is not None else (col_idx + 1)
                val = c.get('value', '')
                    
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                
                # Styles
                font_kwargs = {}
                if c.get('bold'):
                    font_kwargs['bold'] = True
                if c.get('italic'):
                    font_kwargs['italic'] = True
                color = c.get('color')
                if color:
                    hex_color = color.replace('#', '').strip()
                    if re.fullmatch(r'[0-9A-Fa-f]{6}', hex_color):
                        font_kwargs['color'] = "FF" + hex_color.upper()
                    elif re.fullmatch(r'[0-9A-Fa-f]{8}', hex_color):
                        font_kwargs['color'] = hex_color.upper()
                if font_kwargs:
                    cell.font = Font(**font_kwargs)

                bg = c.get('bg')
                if bg:
                    hex_bg = bg.replace('#', '').strip()
                    if re.fullmatch(r'[0-9A-Fa-f]{6}', hex_bg):
                        hex_bg = "FF" + hex_bg.upper()
                    elif re.fullmatch(r'[0-9A-Fa-f]{8}', hex_bg):
                        hex_bg = hex_bg.upper()
                    else:
                        hex_bg = None
                    if hex_bg:
                        cell.fill = PatternFill(start_color=hex_bg, end_color=hex_bg, fill_type="solid")
                
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # Merge cells
                rowspan = c.get('rowspan', 1)
                colspan = c.get('colspan', 1)
                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(
                        start_row=r_idx,
                        start_column=c_idx,
                        end_row=r_idx + rowspan - 1,
                        end_column=c_idx + colspan - 1
                    )
        
        # Autofit column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str:
                    line_lens = [sum(2 if ord(char) > 256 else 1 for char in line) for line in val_str.split('\n')]
                    max_len = max(max_len, max(line_lens))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return Response(
            content=file_stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error("Error generating XLSX: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail="產生 Excel 檔案失敗")


# Mount static files for frontend assets
frontend_dir = Path(__file__).parent.parent / "frontend"
if frontend_dir.exists():
    app.mount("/static", StaticFiles(directory=str(frontend_dir)), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=7860)