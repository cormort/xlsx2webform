#!/usr/bin/env python3
"""
XLSX to HTML Table Converter
Ported from xlsx2html/excel2html.rb to Python + openpyxl.
Handles merged cells, styles, colors, and generates clean HTML tables.
"""

import openpyxl
import json
import re
import logging

logger = logging.getLogger(__name__)


# Color mapping for markers (ported from excel2html.rb)
MARKERS = {
    'F79646': 'marker-orange', 'E46C0A': 'marker-orange', 'FFC000': 'marker-orange',
    'E6B9B8': 'marker-orange', 'D99694': 'marker-orange', 'FF9900': 'marker-orange',
    '996633': 'marker-brown', '984807': 'marker-brown', '948A54': 'marker-brown',
    'CC9900': 'marker-orange', 'CC6600': 'marker-orange',
    '4A452A': 'marker-terra-cota', 'FFFF00': 'marker-yellow',
    '0000FF': 'marker-blue', '0070C0': 'marker-blue', '00B0F0': 'marker-blue',
    '4BACC6': 'marker-blue', '558ED5': 'marker-blue', 'B7DEE8': 'marker-blue',
    '93CDDD': 'marker-blue',
    '31859C': 'marker-dark-blue', '4F81BD': 'marker-dark-blue',
    '1F497D': 'marker-dark-blue', '376092': 'marker-dark-blue',
    '002060': 'marker-dark-blue', '10253F': 'marker-dark-blue',
    '17375E': 'marker-dark-blue', '215968': 'marker-dark-blue',
    '254061': 'marker-dark-blue',
    '008080': 'marker-green', '006600': 'marker-green', '009900': 'marker-green',
    '00B050': 'marker-green', '92D050': 'marker-green', '9BBB59': 'marker-green',
    '77933C': 'marker-green', '4F6228': 'marker-green',
    '6600FF': 'marker-purple', '7030A0': 'marker-purple', '8064A2': 'marker-purple',
    'B3A2C7': 'marker-purple', 'CCC1DA': 'marker-purple', '604A7B': 'marker-purple',
    '9900FF': 'marker-purple', '9933FF': 'marker-purple',
    'FF66CC': 'marker-pink', 'FF00FF': 'marker-pink',
    'C00000': 'marker-red', 'FF0000': 'marker-red', 'C0504D': 'marker-red',
    '953735': 'marker-brique', '632523': 'marker-brique',
    '808080': 'marker-grey', 'A6A6A6': 'marker-grey', 'BFBFBF': 'marker-grey',
    'D9D9D9': 'marker-grey',
}

# Reverse mapping: hex color without # prefix → marker class
COLOR_TO_MARKER = {k.upper(): v for k, v in MARKERS.items()}


import re

_HEX_COLOR_RE = re.compile(r'^[0-9A-Fa-f]{6}(?:[0-9A-Fa-f]{2})?$')


def _sanitize_color(raw):
    """Return a 6-char uppercase hex string, or None if invalid."""
    if not raw:
        return None
    s = str(raw).strip()
    if not _HEX_COLOR_RE.match(s):
        return None
    if len(s) == 8:
        return s[2:].upper()
    return s.upper()


def hex_to_rgb(hex_color):
    """Convert openpyxl Color object or hex string to RGB hex string."""
    if hex_color is None:
        return None

    if isinstance(hex_color, str):
        return _sanitize_color(hex_color)

    if hasattr(hex_color, 'rgb') and hex_color.rgb:
        return _sanitize_color(hex_color.rgb)

    return None


def get_fill_color(cell):
    """Get cell background color as hex string (without #)."""
    fill = cell.fill
    if fill and fill.fgColor:
        color = fill.fgColor
        if color.rgb and str(color.rgb) != '00000000':
            return _sanitize_color(color.rgb)
    return None


def get_font_color(cell):
    """Get cell font color as hex string."""
    if cell.font and cell.font.color:
        color = cell.font.color
        if color.rgb:
            return _sanitize_color(color.rgb)
    return None


def get_marker_class(rgb_hex):
    """Get marker CSS class for a given RGB hex color."""
    if rgb_hex is None:
        return None
    return COLOR_TO_MARKER.get(rgb_hex.upper())


def is_black(color_hex):
    """Check if a color is effectively black."""
    if color_hex is None:
        return True
    return color_hex.upper() in ('000000', 'FF000000', '0D0D0D')


def build_merged_cells_map(ws):
    """Build a map of merged cells: (row, col) → (rowspan, colspan, is_top_left)."""
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        max_row, max_col = merge_range.max_row, merge_range.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    merged[(r, c)] = {
                        'rowspan': rowspan,
                        'colspan': colspan,
                        'is_top_left': True
                    }
                else:
                    merged[(r, c)] = {'skip': True}

    return merged


def cell_to_html(cell, merged_map):
    """Convert a single cell to HTML."""
    row, col = cell.row, cell.column

    # Check if this cell is in a merged range (not top-left)
    if (row, col) in merged_map:
        info = merged_map[(row, col)]
        if info.get('skip'):
            return None
        # Top-left of merged range
        span_attrs = ''
        if info['rowspan'] > 1:
            span_attrs += f" rowspan=\"{info['rowspan']}\""
        if info['colspan'] > 1:
            span_attrs += f" colspan=\"{info['colspan']}\""
    else:
        span_attrs = ''

    # Build style attributes
    style_parts = []

    # Background color
    bg_color = get_fill_color(cell)
    if bg_color and bg_color != 'FFFFFF':
        style_parts.append(f"background-color: #{bg_color};")

    # Font styles
    font = cell.font
    is_bold = font and font.bold
    is_italic = font and font.italic
    is_underline = font and font.underline
    is_strikethrough = font and font.strike

    # Font color
    font_color = get_font_color(cell)
    marker_class = get_marker_class(font_color) if font_color and not is_black(font_color) else None

    # Cell value
    value = cell.value
    if value is None:
        value_str = ''
    else:
        value_str = str(value)

    # Build HTML
    style_attr = f" style=\"{' '.join(style_parts)}\"" if style_parts else ''

    html = f"<td{span_attrs}{style_attr}>"

    # Wrap with font styles
    if marker_class:
        html += f"<mark class=\"{marker_class}\">"

    if is_bold:
        html += '<b>'
    if is_italic:
        html += '<i>'
    if is_underline:
        html += '<u>'
    if is_strikethrough:
        html += '<s>'

    # Escape HTML in value
    escaped = value_str.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    html += escaped

    if is_strikethrough:
        html += '</s>'
    if is_underline:
        html += '</u>'
    if is_italic:
        html += '</i>'
    if is_bold:
        html += '</b>'

    if marker_class:
        html += '</mark>'

    html += '</td>\n'
    return html


def worksheet_to_html(ws, header_rows=0):
    """Convert an openpyxl worksheet to an HTML table string.

    Args:
        ws: openpyxl worksheet
        header_rows: number of header rows to wrap in <thead>

    Returns:
        HTML string with <table> containing the worksheet data
    """
    merged_map = build_merged_cells_map(ws)

    html = '<table>\n'

    # Add table headers if specified
    if header_rows > 0:
        html += '<thead>\n'

    rows_html = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        row_cells = []
        skip_row = False

        for cell in row:
            if (cell.row, cell.column) in merged_map and merged_map[(cell.row, cell.column)].get('skip'):
                continue
            cell_html = cell_to_html(cell, merged_map)
            if cell_html:
                row_cells.append(cell_html)

        if row_cells:
            row_html = '<tr>\n' + ''.join(row_cells) + '</tr>\n'
            rows_html.append((row_idx, row_html))

    if header_rows > 0:
        for _, row_html in rows_html[:header_rows]:
            html += row_html
        html += '</thead>\n<tbody>\n'
        for _, row_html in rows_html[header_rows:]:
            html += row_html
        html += '</tbody>\n'
    else:
        for _, row_html in rows_html:
            html += row_html

    html += '</table>'
    return html


# NOTE: extract_budget_data is currently unused (no route calls it).
# Retained for potential future use.
def extract_budget_data(ws, config=None):
    """Extract structured budget data from a worksheet.

    Args:
        ws: openpyxl worksheet
        config: optional dict with extraction rules

    Returns:
        dict with rows and metadata
    """
    merged_map = build_merged_cells_map(ws)
    rows = []
    headers = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        row_data = []
        for cell in row:
            if (cell.row, cell.column) in merged_map and merged_map[(cell.row, cell.column)].get('skip'):
                row_data.append(None)
                continue

            value = cell.value
            if value is None:
                row_data.append('')
            else:
                row_data.append(str(value))

        if row_idx == 1:
            headers = row_data
        else:
            rows.append(row_data)

    return {
        'headers': headers,
        'rows': rows,
        'row_count': len(rows),
        'col_count': len(headers) if headers else (ws.max_column or 0)
    }


def worksheet_to_json(ws):
    """Convert worksheet to a JSON-serializable structure for the budget editor."""
    merged_map = build_merged_cells_map(ws)

    result = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        row_cells = []
        for cell in row:
            if (cell.row, cell.column) in merged_map and merged_map[(cell.row, cell.column)].get('skip'):
                continue

            value = cell.value
            cell_data = {
                'row': cell.row,
                'col': cell.column,
                'value': str(value) if value is not None else '',
                'type': type(value).__name__ if value is not None else 'NoneType'
            }

            # Add style info
            bg = get_fill_color(cell)
            if bg:
                cell_data['bg'] = bg

            font_color = get_font_color(cell)
            if font_color and not is_black(font_color):
                cell_data['color'] = font_color

            font = cell.font
            if font:
                if font.bold:
                    cell_data['bold'] = True
                if font.italic:
                    cell_data['italic'] = True

            # Add merged cell info
            if (cell.row, cell.column) in merged_map:
                info = merged_map[(cell.row, cell.column)]
                if info.get('rowspan', 1) > 1:
                    cell_data['rowspan'] = info['rowspan']
                if info.get('colspan', 1) > 1:
                    cell_data['colspan'] = info['colspan']

            row_cells.append(cell_data)

        result.append(row_cells)

    return result


def extract_form_data(ws):
    """Extract form structure from worksheet.

    Assumes typical form layout:
    - Column A: labels/field names
    - Column B: values (or empty for user input)
    - Rows with only label in Column A = section headers

    Returns:
        dict with form structure for rendering
    """
    merged_map = build_merged_cells_map(ws)

    sections = []
    current_section = {"label": "", "fields": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column or 2, 10)), start=1):
        # Get first two non-empty cells
        cells = []
        for cell in row[:10]:  # Limit to first 10 columns
            if (cell.row, cell.column) in merged_map and merged_map[(cell.row, cell.column)].get('skip'):
                continue
            cells.append(cell)

        if len(cells) < 1:
            continue

        label_cell = cells[0]
        value_cell = cells[1] if len(cells) > 1 else None

        label = str(label_cell.value).strip() if label_cell.value is not None else ''
        value = str(value_cell.value).strip() if value_cell and value_cell.value is not None else ''

        # Skip completely empty rows
        if not label and not value:
            continue

        # Determine if this is a section header or a field
        # Section header: has label but no value, or label ends with colon/：
        is_section = (
            (label and not value) or
            label.endswith('：') or
            label.endswith(':') or
            (value_cell is None and label)
        )

        if is_section:
            # Save previous section if it has fields
            if current_section['fields']:
                sections.append(current_section)
            # Start new section
            current_section = {"label": label.rstrip('：:'), "fields": []}
        else:
            # This is a field
            # Determine input type based on value
            input_type = "text"
            if value:
                try:
                    float(value.replace(',', ''))
                    input_type = "number"
                except ValueError:
                    input_type = "text"

            # Check if this might be a number field based on label
            if any(kw in label for kw in ['金額', '數量', '預算', '決算', '費用', '收入', '支出', '總計', '合計']):
                input_type = "number"

            current_section['fields'].append({
                "label": label,
                "value": value,
                "type": input_type,
                "row": row_idx
            })

    # Add last section
    if current_section['fields']:
        sections.append(current_section)

    return {
        "title": ws.title or "表單",
        "sections": sections,
        "total_fields": sum(len(s['fields']) for s in sections)
    }


def process_xlsx(file_path, sheet_index=0, mode='table'):
    """Process an XLSX file and return data.

    Args:
        file_path: path to the XLSX file
        sheet_index: which sheet to process (0-based)
        mode: 'table' for table view, 'form' for form view

    Returns:
        dict with 'html', 'json', 'sheet_names', 'metadata'
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        sheet_names = wb.sheetnames
        ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.worksheets[0]

        if mode == 'form':
            form_data = extract_form_data(ws)
            metadata = {
                'sheet_name': ws.title,
                'sheet_names': sheet_names,
                'mode': 'form',
                **{k: v for k, v in form_data.items() if k != 'sections'}
            }
            return {
                'form': form_data,
                'metadata': metadata
            }
        else:
            html = worksheet_to_html(ws)
            json_data = worksheet_to_json(ws)

            metadata = {
                'sheet_name': ws.title,
                'row_count': ws.max_row or 0,
                'col_count': ws.max_column or 0,
                'merged_cells_count': len(ws.merged_cells.ranges),
                'sheet_names': sheet_names,
                'mode': 'table'
            }

            return {
                'html': html,
                'json': json_data,
                'metadata': metadata
            }
    finally:
        try:
            wb.close()
        except Exception:
            pass


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        mode = 'form' if '--form' in sys.argv else 'table'
        result = process_xlsx(sys.argv[1], mode=mode)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print("Usage: python xlsx_parser.py <file.xlsx> [--form]")